from __future__ import annotations

import csv
import hashlib
import re
from zipfile import BadZipFile
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from functools import lru_cache
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.config import PRICING_DATA_MODE, SAP_BASE_CURRENCY


DEFAULT_WORKBOOK_PATH = (
    Path(__file__).resolve().parents[1]
    / "Data"
    / "SAP_archived"
    / "Compass CSI WW.xlsx"
)
SYNTHETIC_PRICING_PATH = (
    Path(__file__).resolve().parents[1]
    / "Data"
    / "synthetic"
    / "pricing_demo.csv"
)

HEADER_ALIASES = {
    "vertical": "vertical",
    "lob": "lob",
    "prod type": "product_type",
    "prod family": "product_family",
    "cat": "product_id",
    "description": "description",
    "prod hierarchy": "product_hierarchy",
    "list price": "list_price",
    "net price": "net_price",
    "cogs": "cogs",
    "installation cogs": "installation_cogs",
    "warranty cogs": "warranty_cogs",
    "cogs i w": "cogs_installation_warranty",
    "cogs mov avg": "cogs_moving_average",
    "freight": "freight",
    "duty": "duty",
    "tariff": "tariff",
    "minimum price": "minimum_price",
    "transfer price": "transfer_price",
    "service net price": "service_net_price",
    "service cogs": "service_cogs",
    "app cogs": "app_cogs",
    "service margin": "service_margin",
    "service margin percent": "service_margin_percent",
}

REQUIRED_COLUMNS = frozenset(
    {
        "product_id",
        "description",
        "list_price",
        "net_price",
        "minimum_price",
        "cogs",
    }
)

MONEY_COLUMNS = (
    "list_price",
    "net_price",
    "minimum_price",
    "transfer_price",
    "cogs",
    "installation_cogs",
    "warranty_cogs",
    "cogs_installation_warranty",
    "cogs_moving_average",
    "freight",
    "duty",
    "tariff",
    "service_net_price",
    "service_cogs",
    "app_cogs",
    "service_margin",
    "service_margin_percent",
)


class PricingDataError(Exception):
    pass


class PricingWorkbookMissingError(PricingDataError):
    pass


class PricingWorkbookSchemaError(PricingDataError):
    pass


@dataclass(frozen=True)
class PricingRecord:
    source_id: str
    source_sheet: str
    product_id: str
    description: str
    product_family: str = ""
    product_type: str = ""
    list_price: Decimal | None = None
    net_price: Decimal | None = None
    minimum_price: Decimal | None = None
    transfer_price: Decimal | None = None
    cogs: Decimal | None = None
    installation_cogs: Decimal | None = None
    warranty_cogs: Decimal | None = None
    cogs_installation_warranty: Decimal | None = None
    cogs_moving_average: Decimal | None = None
    freight: Decimal | None = None
    duty: Decimal | None = None
    tariff: Decimal | None = None
    service_net_price: Decimal | None = None
    service_cogs: Decimal | None = None
    app_cogs: Decimal | None = None
    service_margin: Decimal | None = None
    service_margin_percent: Decimal | None = None
    currency: str = SAP_BASE_CURRENCY


def load_pricing_records(
    source_path: str | Path | None = None,
    *,
    data_mode: str | None = None,
) -> tuple[PricingRecord, ...]:
    mode = (data_mode or PRICING_DATA_MODE).strip().casefold()
    if source_path is not None:
        path = Path(source_path)
    elif mode == "synthetic":
        path = SYNTHETIC_PRICING_PATH
    elif mode == "archived_workbook":
        path = DEFAULT_WORKBOOK_PATH
    else:
        raise PricingDataError(f"Unsupported pricing data mode: {mode}")
    resolved_path = path.resolve()
    if not resolved_path.exists():
        if path.suffix.casefold() == ".csv":
            raise PricingWorkbookMissingError(
                "Synthetic pricing dataset was not found."
            )
        raise PricingWorkbookMissingError(
            "Archived SAP pricing workbook was not found."
        )
    if path.suffix.casefold() == ".csv":
        return _load_pricing_csv_cached(
            str(resolved_path),
            resolved_path.stat().st_mtime_ns,
        )
    return _load_pricing_records_cached(
        str(resolved_path),
        resolved_path.stat().st_mtime_ns,
    )


@lru_cache(maxsize=8)
def _load_pricing_records_cached(
    resolved_path: str,
    modified_time_ns: int,
) -> tuple[PricingRecord, ...]:
    del modified_time_ns
    try:
        workbook = load_workbook(resolved_path, read_only=True, data_only=True)
    except (OSError, BadZipFile, InvalidFileException) as error:
        raise PricingDataError(
            "Archived SAP pricing workbook could not be opened."
        ) from error
    records: list[PricingRecord] = []
    populated_sheets = 0
    try:
        for worksheet in workbook.worksheets:
            sheet_records = tuple(_read_sheet(worksheet))
            if sheet_records:
                populated_sheets += 1
                records.extend(sheet_records)
    finally:
        workbook.close()

    if not populated_sheets:
        raise PricingWorkbookSchemaError(
            "Archived SAP pricing workbook contains no populated sheets."
        )
    return tuple(records)


@lru_cache(maxsize=8)
def _load_pricing_csv_cached(
    resolved_path: str,
    modified_time_ns: int,
) -> tuple[PricingRecord, ...]:
    del modified_time_ns
    try:
        with Path(resolved_path).open(
            "r",
            encoding="utf-8-sig",
            newline="",
        ) as source_file:
            reader = csv.DictReader(source_file)
            headers = reader.fieldnames or []
            canonical_headers = {
                header: HEADER_ALIASES.get(normalize_header(header))
                for header in headers
            }
            available_columns = {
                canonical
                for canonical in canonical_headers.values()
                if canonical
            }
            missing_columns = REQUIRED_COLUMNS.difference(available_columns)
            if missing_columns:
                missing = ", ".join(sorted(missing_columns))
                raise PricingWorkbookSchemaError(
                    f"Synthetic pricing data is missing required columns: {missing}."
                )
            records: list[PricingRecord] = []
            for row_number, row in enumerate(reader, start=2):
                values = {
                    canonical: row.get(header)
                    for header, canonical in canonical_headers.items()
                    if canonical
                }
                if _row_is_blank(tuple(values.values())):
                    continue
                record = _pricing_record_from_values(
                    values,
                    source_name="synthetic_demo",
                    row_number=row_number,
                )
                if record is not None:
                    records.append(record)
    except (OSError, UnicodeError, csv.Error) as error:
        raise PricingDataError(
            "Synthetic pricing data could not be opened."
        ) from error
    if not records:
        raise PricingWorkbookSchemaError(
            "Synthetic pricing data contains no usable records."
        )
    return tuple(records)


def _read_sheet(worksheet: Any) -> Iterable[PricingRecord]:
    header_row_number, column_indexes = _find_header(worksheet)
    if header_row_number is None:
        if _sheet_is_empty(worksheet):
            return ()
        raise PricingWorkbookSchemaError(
            f"Sheet {worksheet.title!r} has no recognizable pricing header."
        )

    missing_columns = REQUIRED_COLUMNS.difference(column_indexes)
    if missing_columns:
        missing = ", ".join(sorted(missing_columns))
        raise PricingWorkbookSchemaError(
            f"Sheet {worksheet.title!r} is missing required columns: {missing}."
        )

    records: list[PricingRecord] = []
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=header_row_number + 1, values_only=True),
        start=header_row_number + 1,
    ):
        if _row_is_blank(row):
            continue
        values = {
            field_name: row[column_index] if column_index < len(row) else None
            for field_name, column_index in column_indexes.items()
        }
        record = _pricing_record_from_values(
            values,
            source_name=worksheet.title,
            row_number=row_number,
        )
        if record is not None:
            records.append(record)
    return tuple(records)


def _pricing_record_from_values(
    values: dict[str, Any],
    *,
    source_name: str,
    row_number: int,
) -> PricingRecord | None:
    product_id = normalize_product_id(values.get("product_id"))
    description = _normalize_text(values.get("description"))
    if not product_id and not description:
        return None
    return PricingRecord(
        source_id=_stable_source_id(source_name, row_number, product_id),
        source_sheet=source_name,
        product_id=product_id,
        description=description,
        product_family=_normalize_text(values.get("product_family")),
        product_type=_normalize_text(values.get("product_type")),
        **{
            field_name: normalize_decimal(values.get(field_name))
            for field_name in MONEY_COLUMNS
        },
    )


def _find_header(worksheet: Any) -> tuple[int | None, dict[str, int]]:
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=20, values_only=True),
        start=1,
    ):
        indexes: dict[str, int] = {}
        for column_index, value in enumerate(row):
            normalized_header = normalize_header(value)
            canonical = HEADER_ALIASES.get(normalized_header)
            if canonical:
                indexes[canonical] = column_index
        if {"product_id", "description"}.issubset(indexes):
            return row_number, indexes
    return None, {}


def normalize_header(value: Any) -> str:
    text = _normalize_text(value).casefold()
    text = text.replace("%", " percent ")
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def normalize_product_id(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if re.fullmatch(r"\d+\.0+", text):
        return text.split(".", 1)[0]
    return text


def normalize_decimal(value: Any) -> Decimal | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text or text in {"-", "--", "N/A", "n/a"}:
            return None
        negative = text.startswith("(") and text.endswith(")")
        text = text.strip("()").replace(",", "").replace("$", "").strip()
        if text.endswith("%"):
            text = text[:-1].strip()
        try:
            result = Decimal(text)
        except InvalidOperation:
            return None
        return -result if negative else result
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _normalize_text(value: Any) -> str:
    return "" if value is None else " ".join(str(value).split())


def _row_is_blank(row: tuple[Any, ...]) -> bool:
    return not any(value is not None and str(value).strip() for value in row)


def _sheet_is_empty(worksheet: Any) -> bool:
    return all(_row_is_blank(row) for row in worksheet.iter_rows(values_only=True))


def _stable_source_id(sheet_name: str, row_number: int, product_id: str) -> str:
    source_key = f"{sheet_name}|{row_number}|{product_id}".encode("utf-8")
    return "SAP-" + hashlib.sha256(source_key).hexdigest()[:16].upper()
