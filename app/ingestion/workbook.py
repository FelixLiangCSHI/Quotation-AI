"""Workbook file validation and sheet reading.

This module is the only place that touches ``openpyxl``. It answers two
questions: *may we open this file at all?* and *what is in it?*
"""

from __future__ import annotations

import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Sequence

from app.ingestion.config import (
    SUPPORTED_EXTENSIONS,
    IngestionConfig,
    load_ingestion_config,
)
from app.ingestion.storage import file_hash

#: OLE2 compound-document signature. Both legacy ``.xls`` files and
#: password-protected OOXML workbooks use this container.
_OLE2_SIGNATURE = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
_ZIP_SIGNATURE = b"PK\x03\x04"

#: Rows scanned when guessing where the header row is.
HEADER_SEARCH_ROWS = 20
#: Rows returned by :func:`read_sheet` previews.
DEFAULT_PREVIEW_ROWS = 25


class WorkbookValidationError(Exception):
    """Raised when an uploaded file cannot be accepted for ingestion."""


class UnsupportedWorkbookError(WorkbookValidationError):
    """The file type is not supported."""


class ProtectedWorkbookError(WorkbookValidationError):
    """The file is encrypted or password protected."""


class CorruptWorkbookError(WorkbookValidationError):
    """The file claims to be a workbook but cannot be parsed."""


@dataclass(frozen=True)
class WorkbookFile:
    """A validated upload that has not yet been parsed into records."""

    filename: str
    extension: str
    size_bytes: int
    content_hash: str
    payload: bytes
    macro_enabled: bool

    @property
    def warnings(self) -> tuple[str, ...]:
        if self.macro_enabled:
            return (
                "The workbook is macro-enabled (.xlsm). Macros are never "
                "executed; only cell values are read.",
            )
        return ()


@dataclass(frozen=True)
class SheetPreview:
    name: str
    header_row: int
    headers: tuple[str, ...]
    rows: tuple[tuple[Any, ...], ...]
    total_rows: int
    #: 1-based worksheet row number for each entry of ``rows``.
    row_numbers: tuple[int, ...] = ()


def validate_workbook_file(
    filename: str,
    payload: bytes,
    *,
    config: IngestionConfig | None = None,
) -> WorkbookFile:
    """Validate an uploaded file before any parsing is attempted."""

    resolved = config or load_ingestion_config()
    name = Path(filename or "").name
    if not name:
        raise UnsupportedWorkbookError("The upload has no file name.")

    extension = Path(name).suffix.casefold()
    if extension not in SUPPORTED_EXTENSIONS:
        supported = ", ".join(SUPPORTED_EXTENSIONS)
        raise UnsupportedWorkbookError(
            f"{name!r} is not a supported workbook. Upload one of: {supported}."
        )

    if not payload:
        raise CorruptWorkbookError(f"{name!r} is empty.")
    if len(payload) > resolved.max_upload_bytes:
        limit_mb = resolved.max_upload_bytes / (1024 * 1024)
        raise UnsupportedWorkbookError(
            f"{name!r} exceeds the {limit_mb:.0f} MB upload limit."
        )

    if payload.startswith(_OLE2_SIGNATURE):
        raise ProtectedWorkbookError(
            f"{name!r} is password protected or is a legacy binary workbook. "
            "Remove the password and re-save it as .xlsx, then upload again."
        )
    if not payload.startswith(_ZIP_SIGNATURE):
        raise CorruptWorkbookError(
            f"{name!r} is not a readable Office Open XML workbook."
        )

    _assert_openable(name, payload, resolved)

    return WorkbookFile(
        filename=name,
        extension=extension,
        size_bytes=len(payload),
        content_hash=file_hash(payload),
        payload=payload,
        macro_enabled=extension == ".xlsm",
    )


def _assert_openable(name: str, payload: bytes, config: IngestionConfig) -> None:
    import io

    try:
        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            infos = archive.infolist()
            entries = {info.filename for info in infos}
    except (zipfile.BadZipFile, OSError) as error:
        raise CorruptWorkbookError(
            f"{name!r} could not be opened as a workbook."
        ) from error

    _assert_not_a_zip_bomb(name, infos, len(payload), config)

    if "EncryptionInfo" in entries or "EncryptedPackage" in entries:
        raise ProtectedWorkbookError(
            f"{name!r} is encrypted. Remove the password and upload again."
        )
    if not any(entry.startswith("xl/") for entry in entries):
        raise CorruptWorkbookError(
            f"{name!r} does not contain a spreadsheet part."
        )


def _assert_not_a_zip_bomb(
    name: str,
    infos: list["zipfile.ZipInfo"],
    compressed_size: int,
    config: IngestionConfig,
) -> None:
    """Refuse archives that would expand far beyond the configured caps.

    An .xlsx file is a ZIP archive, so a small upload can declare gigabytes of
    uncompressed content. The declared sizes are checked before anything is
    read, and no entry is ever extracted to disk.
    """

    if len(infos) > config.max_archive_entries:
        raise UnsupportedWorkbookError(
            f"{name!r} contains too many internal parts to be a valid "
            "pricing workbook."
        )

    uncompressed = 0
    for info in infos:
        entry = info.filename
        if entry.startswith("/") or ".." in Path(entry).parts or ":" in entry:
            raise UnsupportedWorkbookError(
                f"{name!r} contains an unsafe internal path and was rejected."
            )
        uncompressed += max(0, int(info.file_size))
        if uncompressed > config.max_uncompressed_bytes:
            limit_mb = config.max_uncompressed_bytes / (1024 * 1024)
            raise UnsupportedWorkbookError(
                f"{name!r} expands beyond the {limit_mb:.0f} MB decompression "
                "limit and was rejected."
            )

    if compressed_size > 0:
        ratio = uncompressed / compressed_size
        if ratio > config.max_compression_ratio:
            raise UnsupportedWorkbookError(
                f"{name!r} has an implausible compression ratio and was "
                "rejected as a possible decompression bomb."
            )


def _load_workbook(workbook_file: WorkbookFile):
    import io

    try:
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException
    except ImportError as error:  # pragma: no cover - depends on environment
        raise WorkbookValidationError(
            "Reading a workbook requires openpyxl."
        ) from error

    try:
        return load_workbook(
            io.BytesIO(workbook_file.payload),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except (OSError, zipfile.BadZipFile, InvalidFileException, KeyError) as error:
        raise CorruptWorkbookError(
            f"{workbook_file.filename!r} could not be parsed."
        ) from error


def list_sheets(workbook_file: WorkbookFile) -> tuple[str, ...]:
    workbook = _load_workbook(workbook_file)
    try:
        return tuple(workbook.sheetnames)
    finally:
        workbook.close()


def read_sheet(
    workbook_file: WorkbookFile,
    sheet_name: str,
    *,
    header_row: int | None = None,
    max_rows: int | None = None,
    config: IngestionConfig | None = None,
) -> SheetPreview:
    """Read a sheet's header and data rows.

    ``header_row`` is 1-based. When omitted the first row that contains at
    least two non-empty text-like cells is used.
    """

    workbook = _load_workbook(workbook_file)
    try:
        if sheet_name not in workbook.sheetnames:
            raise WorkbookValidationError(
                f"Sheet {sheet_name!r} is not present in the workbook."
            )
        worksheet = workbook[sheet_name]
        row_cap = (config or load_ingestion_config()).max_sheet_rows
        rows = []
        for row in worksheet.iter_rows(values_only=True):
            rows.append(tuple(row))
            if len(rows) > row_cap:
                raise UnsupportedWorkbookError(
                    f"Sheet {sheet_name!r} exceeds the {row_cap} row limit "
                    "for a single pricing sheet."
                )
    finally:
        workbook.close()

    resolved_header = header_row or _guess_header_row(rows)
    if resolved_header is None or resolved_header > len(rows):
        return SheetPreview(
            name=sheet_name,
            header_row=resolved_header or 1,
            headers=(),
            rows=(),
            total_rows=0,
            row_numbers=(),
        )

    headers = tuple(_header_text(value) for value in rows[resolved_header - 1])
    numbered = [
        (resolved_header + 1 + offset, row)
        for offset, row in enumerate(rows[resolved_header:])
        if not row_is_blank(row)
    ]
    total = len(numbered)
    if max_rows is not None:
        numbered = numbered[:max_rows]
    return SheetPreview(
        name=sheet_name,
        header_row=resolved_header,
        headers=headers,
        rows=tuple(row for _, row in numbered),
        total_rows=total,
        row_numbers=tuple(number for number, _ in numbered),
    )


def _guess_header_row(rows: Sequence[tuple[Any, ...]]) -> int | None:
    for index, row in enumerate(rows[:HEADER_SEARCH_ROWS], start=1):
        text_cells = [
            value
            for value in row
            if isinstance(value, str) and value.strip()
        ]
        if len(text_cells) >= 2:
            return index
    return None


def _header_text(value: Any) -> str:
    return "" if value is None else " ".join(str(value).split())


def row_is_blank(row: Sequence[Any]) -> bool:
    return not any(value is not None and str(value).strip() for value in row)
