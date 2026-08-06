"""Minimal, self-contained quotation logic for the on-site demo page.

The demo keeps the existing local natural-language capability (
:mod:`app.natural_language` and :mod:`app.recommender`) and only adds the
pieces the live demo needs: a normalization layer over the parser output,
quotation lines with list and quotation unit prices, a discount-rate approval
rule and in-memory Excel/PDF exports.

Only the discount rate is used for approval. Cost, COGS, margin and profit are
deliberately absent from this module.
"""

from __future__ import annotations

import io
import re
from datetime import date
from typing import Any, Iterable, Mapping, Sequence

#: The single approval rule of the demo.
DISCOUNT_APPROVAL_THRESHOLD = 0.35

#: Synthetic list prices for the demo catalogue, keyed by product id.
DEMO_LIST_PRICES: dict[str, float] = {
    "DEMO-FMT-100": 100000.00,
    "DEMO-FMT-ALT": 82000.00,
    "DEMO-GEN-80": 24000.00,
    "DEMO-STAND-M": 18000.00,
    "DEMO-DETECTOR-43": 15000.00,
    "DEMO-TUBE-400": 12000.00,
    "DEMO-GRID-T": 3500.00,
    "DEMO-TABLE-1": 9000.00,
    "DEMO-MISSING-900": 60000.00,
}

#: Fallback synthetic list prices when a product id is not in the price book.
DEMO_CATEGORY_PRICES: dict[str, float] = {
    "system": 100000.00,
    "generator": 24000.00,
    "tube_stand": 18000.00,
    "detector": 15000.00,
    "tube": 12000.00,
    "grid": 3500.00,
    "table": 9000.00,
    "wallstand": 11000.00,
}

DEFAULT_LIST_PRICE = 20000.00

DEFAULT_CURRENCY = "USD"

#: Regular expressions used to read a discount rate from plain sales language.
_DISCOUNT_PATTERNS = (
    re.compile(r"(\d{1,3}(?:\.\d+)?)\s*%", re.IGNORECASE),
    re.compile(
        r"(\d{1,3}(?:\.\d+)?)\s*(?:percent|per\s*cent|pct)",
        re.IGNORECASE,
    ),
    re.compile(r"折扣\s*[:：]?\s*(\d{1,3}(?:\.\d+)?)", re.IGNORECASE),
)

_DISCOUNT_CONTEXT_RE = re.compile(
    r"discount|off|折扣|优惠|percent|per\s*cent|pct|%",
    re.IGNORECASE,
)

_CUSTOMER_PATTERNS = (
    re.compile(
        r"\b(?:customer|client|account|hospital)\s*(?:is|:|=)\s*"
        r"([A-Z0-9][\w&.\- ]{2,60})",
        re.IGNORECASE,
    ),
    re.compile(r"\bfor\s+([A-Z][\w&.\-]*(?:\s+[A-Z][\w&.\-]*){0,4})\b"),
    re.compile(r"客户\s*[:：]?\s*([^\s,，。]{2,30})"),
)

_REGION_LABELS = {
    "us": "United States",
    "eu": "Europe",
    "china": "China",
    "singapore": "Singapore",
    "malaysia": "Malaysia",
    "indonesia": "Indonesia",
    "canada": "Canada",
}


class DemoQuotationError(ValueError):
    """Raised when a demo quotation cannot be produced."""


def parse_discount_rate(text: str) -> float | None:
    """Read a discount rate such as ``40%`` or ``30 percent`` from ``text``.

    Returns a fraction in ``[0, 1)`` or ``None`` when no rate is stated. The
    parsing is purely local: no AI API is involved.
    """

    if not text:
        return None
    if not _DISCOUNT_CONTEXT_RE.search(text):
        return None
    for pattern in _DISCOUNT_PATTERNS:
        match = pattern.search(text)
        if not match:
            continue
        value = float(match.group(1))
        if 0 <= value < 1:
            # Already expressed as a fraction, e.g. "0.3 discount".
            return value
        if 0 <= value <= 100:
            return value / 100
    return None


def parse_customer_name(text: str) -> str | None:
    """Best-effort local extraction of a customer name."""

    if not text:
        return None
    for pattern in _CUSTOMER_PATTERNS:
        match = pattern.search(text)
        if not match:
            continue
        candidate = match.group(1).strip(" .,:;")
        if candidate and candidate.lower() not in {"the", "a", "an"}:
            return candidate
    return None


def region_label(region: str | None) -> str:
    if not region:
        return ""
    return _REGION_LABELS.get(region.lower(), region.title())


def list_price_for(product_id: str, category: str | None = None) -> float:
    """Synthetic list price of a demo product."""

    if product_id in DEMO_LIST_PRICES:
        return DEMO_LIST_PRICES[product_id]
    if category and category.lower() in DEMO_CATEGORY_PRICES:
        return DEMO_CATEGORY_PRICES[category.lower()]
    return DEFAULT_LIST_PRICE


def empty_configuration() -> dict[str, Any]:
    """The normalized configuration structure used across the demo page."""

    return {
        "customer_name": "",
        "region": "",
        "currency": DEFAULT_CURRENCY,
        "main_product": "",
        "quantity": 1,
        "accessories": [],
        "configuration_description": "",
        "discount_rate": None,
    }


def normalize_configuration(
    request: Any,
    recommendation: Any = None,
    previous: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    """Map the existing parser/recommender output onto one flat structure.

    ``request`` is a :class:`app.natural_language.QuoteRequest` and
    ``recommendation`` a :class:`app.recommender.QuoteRecommendation`. Values
    already known from earlier turns (``previous``) are kept when the new turn
    does not mention them, which is what makes the conversation multi-turn.
    """

    config = dict(previous) if previous else empty_configuration()
    config.setdefault("accessories", [])

    customer = getattr(request, "customer_name", None) or parse_customer_name(
        getattr(request, "raw_text", "") or ""
    )
    if customer:
        config["customer_name"] = customer

    region = region_label(getattr(request, "region", None))
    if region:
        config["region"] = region

    currency = getattr(request, "currency", None)
    if currency:
        config["currency"] = currency.upper()

    quantity = getattr(request, "quantity", None)
    if quantity:
        config["quantity"] = int(quantity)

    discount_rate = parse_discount_rate(getattr(request, "raw_text", "") or "")
    if discount_rate is not None:
        config["discount_rate"] = discount_rate

    main_model = getattr(recommendation, "main_model", None)
    if main_model is not None:
        config["main_product"] = main_model.product_id
        config["main_product_description"] = main_model.short_description
        config["main_product_reason"] = getattr(main_model, "reason", "")

    accessories = getattr(recommendation, "accessories", ()) or ()
    if accessories:
        config["accessories"] = [
            {
                "product_id": item.product_id,
                "description": item.short_description,
                "quantity": int(item.quantity or 1),
            }
            for item in accessories
        ]

    config["configuration_description"] = build_configuration_description(config)
    return config


def build_configuration_description(config: Mapping[str, Any]) -> str:
    """A single human-readable sentence describing the configuration."""

    parts: list[str] = []
    main_description = config.get("main_product_description") or config.get(
        "main_product"
    )
    if main_description:
        parts.append(f"{config.get('quantity', 1)} x {main_description}")
    for accessory in config.get("accessories") or []:
        parts.append(
            f"{accessory.get('quantity', 1)} x {accessory.get('description')}"
        )
    if not parts:
        return ""
    region = config.get("region")
    suffix = f" for {region}" if region else ""
    return "; ".join(parts) + suffix


def missing_configuration_fields(config: Mapping[str, Any]) -> tuple[str, ...]:
    """Fields the assistant still has to ask about."""

    missing: list[str] = []
    if not config.get("main_product"):
        missing.append("main_product")
    if not config.get("customer_name"):
        missing.append("customer_name")
    if not config.get("region"):
        missing.append("region")
    if not config.get("quantity"):
        missing.append("quantity")
    return tuple(missing)


def build_quotation_lines(
    config: Mapping[str, Any],
    discount_rate: float | None = None,
) -> list[dict[str, Any]]:
    """Turn the normalized configuration into editable quotation lines."""

    rate = discount_rate if discount_rate is not None else config.get("discount_rate")
    rate = float(rate or 0.0)
    quantity = int(config.get("quantity") or 1)

    lines: list[dict[str, Any]] = []
    main_product = config.get("main_product")
    if main_product:
        list_unit_price = list_price_for(main_product, "system")
        lines.append(
            {
                "product_code": main_product,
                "description": config.get("main_product_description")
                or main_product,
                "quantity": quantity,
                "list_unit_price": list_unit_price,
                "quotation_unit_price": round(list_unit_price * (1 - rate), 2),
            }
        )
    for accessory in config.get("accessories") or []:
        product_id = accessory.get("product_id")
        if not product_id:
            continue
        list_unit_price = list_price_for(product_id, accessory.get("category"))
        lines.append(
            {
                "product_code": product_id,
                "description": accessory.get("description") or product_id,
                "quantity": int(accessory.get("quantity") or 1) * quantity,
                "list_unit_price": list_unit_price,
                "quotation_unit_price": round(list_unit_price * (1 - rate), 2),
            }
        )
    return [with_line_totals(line) for line in lines]


def with_line_totals(line: Mapping[str, Any]) -> dict[str, Any]:
    """Return ``line`` with its two derived totals recomputed."""

    quantity = max(int(line.get("quantity") or 0), 0)
    list_unit_price = max(float(line.get("list_unit_price") or 0.0), 0.0)
    quotation_unit_price = max(float(line.get("quotation_unit_price") or 0.0), 0.0)
    enriched = dict(line)
    enriched["quantity"] = quantity
    enriched["list_unit_price"] = list_unit_price
    enriched["quotation_unit_price"] = quotation_unit_price
    enriched["list_line_total"] = quantity * list_unit_price
    enriched["quotation_line_total"] = quantity * quotation_unit_price
    return enriched


def compute_totals(lines: Iterable[Mapping[str, Any]]) -> dict[str, float]:
    """List total, quotation total and the resulting discount rate."""

    priced = [with_line_totals(line) for line in lines]
    list_total = sum(line["list_line_total"] for line in priced)
    quotation_total = sum(line["quotation_line_total"] for line in priced)
    discount_rate = (
        (list_total - quotation_total) / list_total if list_total > 0 else 0.0
    )
    return {
        "list_total": list_total,
        "quotation_total": quotation_total,
        "discount_rate": discount_rate,
    }


def approval_status(discount_rate: float) -> str:
    """The only approval rule of this demo.

    # The 35% boundary is included in Sales authority.
    # Only a discount strictly greater than 35% requires manager approval.
    """

    if discount_rate <= DISCOUNT_APPROVAL_THRESHOLD:
        return "AUTO_APPROVED"
    return "MANAGER_APPROVAL_REQUIRED"


def format_money(currency: str, amount: float) -> str:
    return f"{currency or DEFAULT_CURRENCY} {amount:,.2f}"


def build_approval_description(
    quotation_id: str,
    config: Mapping[str, Any],
    totals: Mapping[str, float],
) -> str:
    """The copy-ready approval text shown when the discount exceeds 35%."""

    currency = config.get("currency") or DEFAULT_CURRENCY
    return (
        f"Quotation {quotation_id} for "
        f"{config.get('customer_name') or 'the customer'} requires manager "
        "approval.\n\n"
        "Customer:\n"
        f"{config.get('customer_name') or '-'}\n\n"
        "Region:\n"
        f"{config.get('region') or '-'}\n\n"
        "Configuration:\n"
        f"{config.get('configuration_description') or '-'}\n\n"
        "Currency:\n"
        f"{currency}\n\n"
        "List total:\n"
        f"{currency} {totals['list_total']:,.2f}\n\n"
        "Quotation total:\n"
        f"{currency} {totals['quotation_total']:,.2f}\n\n"
        "Discount rate:\n"
        f"{totals['discount_rate']:.1%}\n\n"
        "Approval threshold:\n"
        f"{DISCOUNT_APPROVAL_THRESHOLD:.1%}\n\n"
        "Reason for approval:\n"
        "The proposed discount rate exceeds the "
        f"{DISCOUNT_APPROVAL_THRESHOLD:.0%} Sales approval authority.\n\n"
        "Please review the attached quotation and approve, reject or request "
        "revision."
    )


def build_quotation_excel(
    quotation_id: str,
    config: Mapping[str, Any],
    lines: Sequence[Mapping[str, Any]],
    totals: Mapping[str, float],
    internal: bool = False,
) -> bytes:
    """Build the quotation workbook in memory (no file is written to disk)."""

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    currency = config.get("currency") or DEFAULT_CURRENCY
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Approval" if internal else "Quotation"

    title = (
        "Internal Approval Quotation" if internal else "Customer Quotation"
    )
    sheet["A1"] = title
    sheet["A1"].font = Font(bold=True, size=14)

    header_rows = [
        ("Quotation ID", quotation_id),
        ("Date", date.today().isoformat()),
        ("Customer", config.get("customer_name") or "-"),
        ("Region", config.get("region") or "-"),
        ("Currency", currency),
        ("Configuration", config.get("configuration_description") or "-"),
    ]
    row = 3
    for label, value in header_rows:
        sheet.cell(row=row, column=1, value=label).font = Font(bold=True)
        sheet.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    columns = [
        "Product Code",
        "Description",
        "Quantity",
        "List Unit Price",
        "Quotation Unit Price",
        "List Total",
        "Quotation Total",
    ]
    for index, column in enumerate(columns, start=1):
        cell = sheet.cell(row=row, column=index, value=column)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    row += 1

    for line in (with_line_totals(item) for item in lines):
        sheet.cell(row=row, column=1, value=line["product_code"])
        sheet.cell(row=row, column=2, value=line["description"])
        sheet.cell(row=row, column=3, value=line["quantity"])
        sheet.cell(row=row, column=4, value=line["list_unit_price"])
        sheet.cell(row=row, column=5, value=line["quotation_unit_price"])
        sheet.cell(row=row, column=6, value=line["list_line_total"])
        sheet.cell(row=row, column=7, value=line["quotation_line_total"])
        row += 1

    row += 1
    summary = [
        ("List Total", totals["list_total"]),
        ("Quotation Total", totals["quotation_total"]),
        ("Discount Rate", f"{totals['discount_rate']:.1%}"),
    ]
    if internal:
        summary.extend(
            [
                (
                    "Approval Threshold",
                    f"{DISCOUNT_APPROVAL_THRESHOLD:.1%}",
                ),
                ("Approval Status", "Manager approval required"),
                (
                    "Reason",
                    "The discount rate exceeds the "
                    f"{DISCOUNT_APPROVAL_THRESHOLD:.0%} Sales approval "
                    "authority.",
                ),
            ]
        )
    for label, value in summary:
        sheet.cell(row=row, column=1, value=label).font = Font(bold=True)
        sheet.cell(row=row, column=2, value=value)
        row += 1

    widths = {"A": 22, "B": 46, "C": 10, "D": 18, "E": 22, "F": 16, "G": 18}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_customer_pdf(
    quotation_id: str,
    config: Mapping[str, Any],
    lines: Sequence[Mapping[str, Any]],
    totals: Mapping[str, float],
) -> bytes:
    """Build the customer PDF in memory using the bundled ReportLab engine."""

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    currency = config.get("currency") or DEFAULT_CURRENCY
    styles = getSampleStyleSheet()
    buffer = io.BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        title=f"Quotation {quotation_id}",
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
    )

    story: list[Any] = [
        Paragraph("Customer Quotation", styles["Title"]),
        Spacer(1, 6 * mm),
        Paragraph(f"Quotation ID: {quotation_id}", styles["Normal"]),
        Paragraph(f"Date: {date.today().isoformat()}", styles["Normal"]),
        Paragraph(
            f"Customer: {config.get('customer_name') or '-'}", styles["Normal"]
        ),
        Paragraph(f"Region: {config.get('region') or '-'}", styles["Normal"]),
        Paragraph(f"Currency: {currency}", styles["Normal"]),
        Spacer(1, 4 * mm),
        Paragraph(
            f"Configuration: {config.get('configuration_description') or '-'}",
            styles["Normal"],
        ),
        Spacer(1, 6 * mm),
    ]

    table_data: list[list[str]] = [
        [
            "Product Code",
            "Description",
            "Qty",
            "List Unit Price",
            "Quotation Unit Price",
            "Quotation Total",
        ]
    ]
    for line in (with_line_totals(item) for item in lines):
        table_data.append(
            [
                str(line["product_code"]),
                str(line["description"]),
                str(line["quantity"]),
                f"{line['list_unit_price']:,.2f}",
                f"{line['quotation_unit_price']:,.2f}",
                f"{line['quotation_line_total']:,.2f}",
            ]
        )

    table = Table(table_data, repeatRows=1, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5F5")),
                ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
            ]
        )
    )
    story.append(table)
    story.append(Spacer(1, 6 * mm))
    story.append(
        Paragraph(
            f"List total: {format_money(currency, totals['list_total'])}",
            styles["Normal"],
        )
    )
    story.append(
        Paragraph(
            "Quotation total: "
            f"{format_money(currency, totals['quotation_total'])}",
            styles["Normal"],
        )
    )
    story.append(
        Paragraph(
            f"Discount rate: {totals['discount_rate']:.1%}", styles["Normal"]
        )
    )
    story.append(Spacer(1, 6 * mm))
    story.append(
        Paragraph(
            "Demo document generated from synthetic data.", styles["Italic"]
        )
    )

    document.build(story)
    return buffer.getvalue()
